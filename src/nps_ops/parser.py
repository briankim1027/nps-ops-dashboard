"""Parser for the 2026-06 NPS operating Excel format.

This module intentionally targets the real operating workbook shape such as
`●26년06월 NPS평가 통계_0622.xlsx`, not the legacy `NPS_v1.xlsx` sample.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REQUIRED_SHEETS = {
    "store": "매장별",
    "agency": "대리점별",
    "team": "■팀별",
    "crew": "T크루별",
    "store_crew": "T매장크루별",
    "negative": "응답_비추천",
    "response": "AI만족도조사_리스트",
    "trend": "일별트렌드",
}

HEADER_ROWS = {
    "매장별": 10,              # Excel row 11
    "대리점별": 9,             # Excel row 10
    "T크루별": 10,             # Excel row 11
    "T매장크루별": 10,          # Excel row 11
    "응답_비추천": 4,           # Excel row 5
    "AI만족도조사_리스트": 4,    # Excel row 5
}


@dataclass(frozen=True)
class WorkbookProfile:
    path: Path
    report_date: date | None
    sheets: dict[str, tuple[int, int]]
    missing_required_sheets: list[str]


def _clean_col(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _dedupe_columns(cols: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in cols:
        base = _clean_col(raw) or "unnamed"
        if base in seen:
            seen[base] += 1
            out.append(f"{base}.{seen[base]}")
        else:
            seen[base] = 0
            out.append(base)
    return out


def _read_table(path: Path, sheet_name: str, header: int) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, header=header)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df.columns = _dedupe_columns(list(df.columns))
    return df


def _to_yyyymmdd(value: Any) -> pd.Series | pd.Timestamp | Any:
    return pd.to_datetime(value, format="%Y%m%d", errors="coerce")


def extract_report_date(path: Path) -> date | None:
    """Extract report date from known cells, falling back to filename suffix."""
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: list[Any] = []
    for sheet, cells in {
        "매장별": ["X1", "H4"],
        "대리점별": ["J1"],
        "■팀별": ["M2"],
        "출력_대리점": ["N9"],
    }.items():
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in cells:
                candidates.append(ws[cell].value)
    for v in candidates:
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        ts = pd.to_datetime(v, errors="coerce")
        if not pd.isna(ts):
            return ts.date()
    m = re.search(r"(?:raw|only|v)_(\d{6})", path.stem, flags=re.IGNORECASE)
    if m:
        yymmdd = m.group(1)
        return date(2000 + int(yymmdd[:2]), int(yymmdd[2:4]), int(yymmdd[4:]))
    m = re.search(r"_(\d{4})(?!\d)", path.stem)
    if m:
        mmdd = m.group(1)
        # Operating file is 2026년06월 for this project scope.
        return date(2026, int(mmdd[:2]), int(mmdd[2:]))
    return None


def profile_workbook(path: str | Path) -> WorkbookProfile:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
    missing = [name for name in REQUIRED_SHEETS.values() if name not in sheets]
    return WorkbookProfile(path=path, report_date=extract_report_date(path), sheets=sheets, missing_required_sheets=missing)


def parse_response_fact(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_table(path, "AI만족도조사_리스트", HEADER_ROWS["AI만족도조사_리스트"])
    rename = {
        "매장코드": "store_code",
        "매장&직원": "store_crew_key",
        "팀명": "team_name",
        "구분": "response_type",
        "추천": "promoter_flag",
        "중립": "passive_flag",
        "비추천": "detractor_flag",
        "업무처리일자": "process_date",
        "평가일자": "evaluation_date",
        "평가대상여부": "evaluation_target_flag",
        "의견있음": "has_comment",
        "처리직원": "staff_name",
        "본부": "hq_name",
        "마케팅팀": "marketing_team",
        "미케팅팀명": "marketing_team_name_raw",
        "대리점": "agency_code",
        "대리점명": "agency_name",
        "매장코드.1": "store_code_lookup",
        "매장명": "store_name",
        "담당자": "marketer",
        "담당자ID": "marketer_id",
        "처리일": "process_date_text",
        "업무유형": "business_type",
        "추천지수": "recommend_score",
        "추천지수확인": "recommend_score_confirmed",
        "추천사유": "comment_text",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    for col in ["promoter_flag", "passive_flag", "detractor_flag", "evaluation_target_flag", "has_comment"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype("int64")
    for col in ["process_date", "evaluation_date"]:
        if col in df.columns:
            df[col] = _to_yyyymmdd(df[col])
    if "store_code_lookup" in df.columns:
        df["store_code"] = df["store_code"].fillna(df["store_code_lookup"])
    if "response_type" in df.columns:
        df["response_type"] = df["response_type"].astype(str).str.strip()
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    return df


def parse_store_agg(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_table(path, "매장별", HEADER_ROWS["매장별"])
    rename = {
        "권역": "region",
        "팀명": "team_name",
        "코드": "agency_code",
        "대리점명": "agency_name",
        "매장코드": "store_code",
        "매장명": "store_name",
        "마케터": "marketer",
        "H/S판매": "hs_sales",
        "발송대상 ①": "message_target",
        "베스트톡 발송 ②": "besttalk_sent",
        "발송율 ②÷①": "besttalk_send_rate",
        "NPS": "prev_nps_score",
        "추천": "prev_promoters",
        "중립": "prev_passives",
        "비추천": "prev_detractors",
        "총응답자": "prev_total_responses",
        "NPS.1": "nps_score",
        "전일대비": "day_delta",
        "추천.1": "promoters",
        "중립.1": "passives",
        "비추천.1": "detractors",
        "총응답자.1": "total_responses",
        "0.9": "required_promoters_90",
        "0.93": "required_promoters_93",
        "5건이상": "five_plus_sample",
        "Grade": "grade",
        "NPS.2": "hq_nps_score",
        "NPS.3": "sales_nps_score",
        "추천.2": "sales_promoters",
        "중립.2": "sales_passives",
        "비추천.2": "sales_detractors",
        "총응답자.2": "sales_total_responses",
        "NPS.4": "non_sales_nps_score",
        "추천.3": "non_sales_promoters",
        "중립.3": "non_sales_passives",
        "비추천.3": "non_sales_detractors",
        "총응답자.3": "non_sales_total_responses",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    # Keep only actual store rows; summary rows have blank region/team mismatch or no store code.
    if "store_code" in df.columns:
        df = df[df["store_code"].notna()].copy()
    for col in df.columns:
        if any(token in col for token in ["nps", "promoters", "passives", "detractors", "responses", "sales", "target", "sent", "rate", "delta"]):
            if col not in {"agency_code", "store_code", "store_name", "team_name", "agency_name", "marketer", "source_file"}:
                df[col] = pd.to_numeric(df[col], errors="coerce")
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    return df


def parse_agency_agg(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_table(path, "대리점별", HEADER_ROWS["대리점별"])
    rename = {
        "팀명": "team_name",
        "대리점코드": "agency_code",
        "대리점명": "agency_name",
        "마케터": "marketer",
        "종합NPS": "nps_score",
        "추천": "promoters",
        "중립": "passives",
        "비추천": "detractors",
        "총응답자": "total_responses",
        "종합NPS.1": "sales_nps_score",
        "추천.1": "sales_promoters",
        "중립.1": "sales_passives",
        "비추천.1": "sales_detractors",
        "총응답자.1": "sales_total_responses",
        "종합NPS.2": "non_sales_nps_score",
        "추천.2": "non_sales_promoters",
        "중립.2": "non_sales_passives",
        "비추천.2": "non_sales_detractors",
        "총응답자.2": "non_sales_total_responses",
        "GAP": "gap",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "agency_code" in df.columns:
        df = df[df["agency_code"].notna()].copy()
    for col in [c for c in df.columns if c not in {"team_name", "agency_code", "agency_name", "marketer", "source_file"}]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    return df


def parse_negative_feedback(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_table(path, "응답_비추천", HEADER_ROWS["응답_비추천"])
    rename = {
        "팀명": "team_name",
        "구분": "response_type",
        "매장코드": "store_code",
        "매장명": "store_name",
        "응답일자": "response_date",
        "처리일자": "process_date",
        "업무유형": "business_type",
        "처리직원": "staff_name",
        "지인추천": "recommend_score",
        "추천의사사유": "reason_text",
        "코드": "agency_code",
        "대리점명": "agency_name",
        "추천사유병합": "reason_merged",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    for col in ["response_date", "process_date"]:
        if col in df.columns:
            df[col] = _to_yyyymmdd(df[col])
    if "recommend_score" in df.columns:
        df["recommend_score"] = pd.to_numeric(df["recommend_score"], errors="coerce")
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    return df


def parse_crew_agg(path: str | Path, sheet_name: str = "T크루별") -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_table(path, sheet_name, HEADER_ROWS[sheet_name])
    rename = {
        "팀명": "team_name",
        "대리점코드": "agency_code",
        "대리점명": "agency_name",
        "매장코드": "store_code",
        "매장명": "store_name",
        "LOGIN-ID": "crew_id",
        "직원명": "crew_name",
        "매장NPS": "store_nps_score",
        "매장총샘플": "store_total_responses",
        "NPS": "nps_score",
        "추천": "promoters",
        "중립": "passives",
        "비추천": "detractors",
        "총응답자": "total_responses",
        "추천지수평균": "avg_recommend_score",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "crew_id" in df.columns:
        df = df[df["crew_id"].notna()].copy()
    for col in [c for c in df.columns if c not in {"team_name", "agency_code", "agency_name", "store_code", "store_name", "crew_id", "crew_name", "source_file", "source_sheet"}]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    df["source_sheet"] = sheet_name
    return df



RAW_LEDGER_REQUIRED_COLUMNS = {"매장코드", "업무처리일자", "평가월", "판매성YN", "본부", "대리점", "대리점명", "매장명", "담당자", "담당자ID", "업무유형", "추천지수", "추천사유"}


def is_raw_ledger_workbook(path: str | Path) -> bool:
    """Return True for the 2026 1~6월 NPS raw ledger workbook shape.

    The raw ledger has one flat sheet with row-level store/agency/staff/date/VOC
    fields, unlike the operating workbook with separate 매장별/대리점별/T크루별 sheets.
    """
    path = Path(path)
    try:
        xls = pd.ExcelFile(path)
        if not xls.sheet_names:
            return False
        cols = set(pd.read_excel(path, sheet_name=xls.sheet_names[0], nrows=0).columns.astype(str))
        return RAW_LEDGER_REQUIRED_COLUMNS.issubset(cols)
    except Exception:
        return False


def _read_raw_ledger(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    xls = pd.ExcelFile(path)
    df = pd.read_excel(
        path,
        sheet_name=xls.sheet_names[0],
        dtype={"매장코드": str, "마케팅팀": str, "대리점": str, "매장": str, "담당자ID": str},
    )
    df = df.dropna(axis=0, how="all").copy()
    df.columns = _dedupe_columns(list(df.columns))
    return df


def parse_raw_response_fact(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    report_date = extract_report_date(path)
    df = _read_raw_ledger(path)
    rename = {
        "매장코드": "store_code",
        "구분": "response_type_raw",
        "업무처리일자": "process_date",
        "평가월": "evaluation_month",
        "판매성YN": "NCSI",
        "본부": "hq_name",
        "마케팅팀": "marketing_team",
        "미케팅팀명": "team_name",
        "마케팅팀명": "team_name",
        "대리점": "agency_code",
        "대리점명": "agency_name",
        "매장": "store_sub_code",
        "매장명": "store_name",
        "담당자": "marketer",
        "담당자ID": "marketer_id",
        "처리일": "process_date_text",
        "업무유형": "business_type",
        "추천지수": "recommend_score",
        "추천지수확인": "recommend_score_confirmed",
        "추천사유": "comment_text",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "process_date" in df.columns:
        df["process_date"] = _to_yyyymmdd(df["process_date"].astype(str).str.replace(r"\.0$", "", regex=True))
    if "recommend_score" in df.columns:
        df["recommend_score"] = pd.to_numeric(df["recommend_score"], errors="coerce")
    score = pd.to_numeric(df.get("recommend_score", pd.Series(index=df.index, dtype="float64")), errors="coerce")
    df["promoter_flag"] = score.ge(9).fillna(False).astype("int64")
    df["passive_flag"] = score.between(7, 8, inclusive="both").fillna(False).astype("int64")
    df["detractor_flag"] = score.le(6).fillna(False).astype("int64")
    if "team_name" in df.columns:
        df["team_name_full"] = df["team_name"].astype(str).str.strip()
        df["team_name"] = df["team_name_full"].str.replace("마케팅팀", "", regex=False).str.strip()
    df["response_type"] = pd.Series("", index=df.index, dtype="object")
    df.loc[df["promoter_flag"].eq(1), "response_type"] = "추천"
    df.loc[df["passive_flag"].eq(1), "response_type"] = "중립"
    df.loc[df["detractor_flag"].eq(1), "response_type"] = "비추천"
    df["evaluation_date"] = pd.NaT
    df["report_date"] = pd.Timestamp(report_date) if report_date else pd.NaT
    df["source_file"] = path.name
    return df


def _nps_from_group(g: pd.DataFrame) -> float | None:
    total = len(g)
    if total == 0:
        return None
    return (int(g["promoter_flag"].sum()) - int(g["detractor_flag"].sum())) / total * 100


def parse_raw_store_agg(path: str | Path) -> pd.DataFrame:
    rf = parse_raw_response_fact(path)
    id_cols = ["team_name", "agency_code", "agency_name", "store_code", "store_name", "marketer"]
    for c in id_cols:
        if c not in rf.columns:
            rf[c] = pd.NA
    base = (
        rf.groupby(id_cols, dropna=False)
        .agg(promoters=("promoter_flag", "sum"), passives=("passive_flag", "sum"), detractors=("detractor_flag", "sum"), total_responses=("promoter_flag", "size"))
        .reset_index()
    )
    axes = []
    for axis_value, prefix in [("판매성", "sales_"), ("비판매성", "non_sales_")]:
        part = rf[rf.get("NCSI", "").astype(str).str.strip().eq(axis_value)].copy()
        agg = (
            part.groupby(id_cols, dropna=False)
            .agg(**{f"{prefix}promoters": ("promoter_flag", "sum"), f"{prefix}passives": ("passive_flag", "sum"), f"{prefix}detractors": ("detractor_flag", "sum"), f"{prefix}total_responses": ("promoter_flag", "size")})
            .reset_index()
        ) if not part.empty else pd.DataFrame(columns=id_cols + [f"{prefix}promoters", f"{prefix}passives", f"{prefix}detractors", f"{prefix}total_responses"])
        axes.append(agg)
    out = base.copy()
    for agg in axes:
        out = out.merge(agg, on=id_cols, how="left")
    count_cols = [c for c in out.columns if c.endswith(("promoters", "passives", "detractors", "responses"))]
    for c in count_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
    out["nps_score"] = out.apply(lambda r: (r.promoters - r.detractors) / r.total_responses * 100 if r.total_responses else None, axis=1)
    out["sales_nps_score"] = out.apply(lambda r: (r.sales_promoters - r.sales_detractors) / r.sales_total_responses * 100 if r.sales_total_responses else None, axis=1)
    out["non_sales_nps_score"] = out.apply(lambda r: (r.non_sales_promoters - r.non_sales_detractors) / r.non_sales_total_responses * 100 if r.non_sales_total_responses else None, axis=1)
    out["prev_nps_score"] = pd.NA
    out["hq_nps_score"] = pd.NA
    out["report_date"] = rf["report_date"].iloc[0] if "report_date" in rf.columns and len(rf) else pd.NaT
    out["source_file"] = Path(path).name
    return out


def parse_raw_agency_agg(path: str | Path) -> pd.DataFrame:
    rf = parse_raw_response_fact(path)
    id_cols = ["team_name", "agency_code", "agency_name", "marketer"]
    for c in id_cols:
        if c not in rf.columns:
            rf[c] = pd.NA
    out = (
        rf.groupby(id_cols, dropna=False)
        .agg(promoters=("promoter_flag", "sum"), passives=("passive_flag", "sum"), detractors=("detractor_flag", "sum"), total_responses=("promoter_flag", "size"))
        .reset_index()
    )
    for axis_value, prefix in [("판매성", "sales_"), ("비판매성", "non_sales_")]:
        part = rf[rf.get("NCSI", "").astype(str).str.strip().eq(axis_value)].copy()
        agg = (
            part.groupby(id_cols, dropna=False)
            .agg(**{f"{prefix}promoters": ("promoter_flag", "sum"), f"{prefix}passives": ("passive_flag", "sum"), f"{prefix}detractors": ("detractor_flag", "sum"), f"{prefix}total_responses": ("promoter_flag", "size")})
            .reset_index()
        ) if not part.empty else pd.DataFrame(columns=id_cols + [f"{prefix}promoters", f"{prefix}passives", f"{prefix}detractors", f"{prefix}total_responses"])
        out = out.merge(agg, on=id_cols, how="left")
    for c in [c for c in out.columns if c.endswith(("promoters", "passives", "detractors", "responses"))]:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
    out["nps_score"] = out.apply(lambda r: (r.promoters - r.detractors) / r.total_responses * 100 if r.total_responses else None, axis=1)
    out["sales_nps_score"] = out.apply(lambda r: (r.sales_promoters - r.sales_detractors) / r.sales_total_responses * 100 if r.sales_total_responses else None, axis=1)
    out["non_sales_nps_score"] = out.apply(lambda r: (r.non_sales_promoters - r.non_sales_detractors) / r.non_sales_total_responses * 100 if r.non_sales_total_responses else None, axis=1)
    out["report_date"] = rf["report_date"].iloc[0] if "report_date" in rf.columns and len(rf) else pd.NaT
    out["source_file"] = Path(path).name
    return out


def parse_raw_negative_feedback(path: str | Path) -> pd.DataFrame:
    rf = parse_raw_response_fact(path)
    neg = rf[pd.to_numeric(rf.get("recommend_score"), errors="coerce").le(8)].copy()
    rename = {"comment_text": "reason_text", "process_date": "response_date"}
    neg = neg.rename(columns={k: v for k, v in rename.items() if k in neg.columns})
    if "response_date" in neg.columns:
        neg["process_date"] = neg["response_date"]
    return neg


def parse_raw_crew_agg(path: str | Path, by_store: bool = False) -> pd.DataFrame:
    rf = parse_raw_response_fact(path)
    id_cols = ["team_name", "agency_code", "agency_name", "marketer_id", "marketer"]
    if by_store:
        id_cols = ["team_name", "agency_code", "agency_name", "store_code", "store_name", "marketer_id", "marketer"]
    for c in id_cols:
        if c not in rf.columns:
            rf[c] = pd.NA
    out = (
        rf.groupby(id_cols, dropna=False)
        .agg(promoters=("promoter_flag", "sum"), passives=("passive_flag", "sum"), detractors=("detractor_flag", "sum"), total_responses=("promoter_flag", "size"), avg_recommend_score=("recommend_score", "mean"))
        .reset_index()
    )
    out = out.rename(columns={"marketer_id": "crew_id", "marketer": "crew_name"})
    out["nps_score"] = out.apply(lambda r: (r.promoters - r.detractors) / r.total_responses * 100 if r.total_responses else None, axis=1)
    out["report_date"] = rf["report_date"].iloc[0] if "report_date" in rf.columns and len(rf) else pd.NaT
    out["source_file"] = Path(path).name
    out["source_sheet"] = "raw_ledger_store_crew" if by_store else "raw_ledger_crew"
    return out

def parse_workbook(path: str | Path) -> dict[str, pd.DataFrame]:
    path = Path(path)
    if is_raw_ledger_workbook(path):
        return {
            "response_fact": parse_raw_response_fact(path),
            "store_agg": parse_raw_store_agg(path),
            "agency_agg": parse_raw_agency_agg(path),
            "negative_feedback": parse_raw_negative_feedback(path),
            "crew_agg": parse_raw_crew_agg(path, by_store=False),
            "store_crew_agg": parse_raw_crew_agg(path, by_store=True),
        }
    return {
        "response_fact": parse_response_fact(path),
        "store_agg": parse_store_agg(path),
        "agency_agg": parse_agency_agg(path),
        "negative_feedback": parse_negative_feedback(path),
        "crew_agg": parse_crew_agg(path, "T크루별"),
        "store_crew_agg": parse_crew_agg(path, "T매장크루별"),
    }
